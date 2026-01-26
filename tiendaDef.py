import tkinter as tk
from tkinter import messagebox
from baseSystem import BaseSystem
import styles
from datetime import datetime
import os

class TiendaSystem(BaseSystem):
    def __init__(self, root, return_callback, db_status="Conectado"):
        # Configurar botón personalizado para el encabezado
        custom_button = {
            'text': "EXPORTAR EXCEL",
            'command': self.exportar_excel_tienda
        }
        
        # Llamar al constructor de BaseSystem con el botón personalizado
        super().__init__(root, return_callback, "TIENDA", db_status, custom_button)
        
        # Actualizar alarmas al iniciar
        self.update_alarms()
    
    def open_productos(self):
        """Abre la ventana de gestión de productos TIENDA - SOBRESCRIBE método de BaseSystem"""
        try:
            from ventanaProductosTienda import VentanaProductosTienda
            VentanaProductosTienda(self.root, self.system_name)
        except ImportError as e:
            messagebox.showerror("Error", f"No se pudo abrir gestión de productos: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir productos TIENDA: {e}")
    
    def get_alarmas(self):
        """Obtiene las alarmas específicas de TIENDA desde la base de datos - SOBRESCRIBE método de BaseSystem"""
        try:
            from db import Database
            db = Database()
            return db.get_alarmas_tienda()
        except Exception as e:
            print(f"Error obteniendo alarmas TIENDA: {e}")
            return None
    
    def show_alarm_details(self):
        """Abre ventana para ver detalles de alarmas TIENDA - SOBRESCRIBE método de BaseSystem"""
        try:
            from ventanaAlarmasTienda import VentanaAlarmasTienda
            VentanaAlarmasTienda(self.root, self.system_name)
        except ImportError as e:
            print(f"ventanaAlarmasTienda no encontrada: {e}")
            # Si no existe ventana específica, abrir productos
            self.open_productos()
    
    def exportar_excel_tienda(self):
        """Exporta datos de TIENDA a Excel con confirmación previa"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from db import Database
            from datetime import datetime
            import os
            
            # Primero obtener datos para mostrar en la confirmación
            db = Database()
            data = db.get_export_data_tienda()
            
            if not data['resumen'] and not data['detalle']:
                messagebox.showwarning("Sin datos", "No hay datos para exportar en TIENDA")
                return
            
            # Calcular estadísticas
            total_productos = len(data['resumen'])
            total_registros = len(data['detalle'])
            
            # Crear ventana de confirmación personalizada (MODAL)
            confirm_window = tk.Toplevel(self.root)
            confirm_window.title("Confirmar Exportación")
            confirm_window.geometry("550x350")
            confirm_window.configure(bg=styles.COLOR_FONDO)
            confirm_window.resizable(False, False)
            
            # Hacerla modal
            confirm_window.transient(self.root)
            confirm_window.grab_set()
            
            # Centrar sobre la ventana principal
            self.root.update_idletasks()
            root_x = self.root.winfo_x()
            root_y = self.root.winfo_y()
            root_width = self.root.winfo_width()
            root_height = self.root.winfo_height()
            
            window_width = 550
            window_height = 350
            
            x = root_x + (root_width // 2) - (window_width // 2)
            y = root_y + (root_height // 2) - (window_height // 2)
            
            confirm_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
            
            # Frame principal
            main_frame = tk.Frame(confirm_window, 
                                bg=styles.COLOR_FONDO, 
                                padx=30, 
                                pady=25)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Icono/emoji principal (EXPORTAR)
            tk.Label(main_frame, 
                    text="📥",
                    font=("Segoe UI Emoji", 32),
                    bg=styles.COLOR_FONDO,
                    fg=self.system_color).pack(pady=(0, 10))
            
            # Título
            tk.Label(main_frame,
                    text="Exportar a Excel",
                    font=(styles.FUENTE_PRINCIPAL, 16, styles.PESO_NEGRITA),
                    bg=styles.COLOR_FONDO,
                    fg=styles.COLOR_TEXTO_OSCURO).pack(pady=(0, 8))
            
            # Mensaje descriptivo
            tk.Label(main_frame,
                    text="Se exportarán los siguientes datos:",
                    font=(styles.FUENTE_PRINCIPAL, 11),
                    bg=styles.COLOR_FONDO,
                    fg=styles.COLOR_TEXTO_MEDIO).pack(pady=(0, 20))
            
            # CONTENEDOR CON ICONO GRANDE Y ESTADÍSTICAS AL LADO
            stats_container = tk.Frame(main_frame, bg=styles.COLOR_FONDO)
            stats_container.pack(fill=tk.X, pady=(0, 25))
            
            # LADO IZQUIERDO: ICONO GRANDE DE GRÁFICO
            icon_frame = tk.Frame(stats_container, bg=styles.COLOR_FONDO)
            icon_frame.pack(side=tk.LEFT, padx=(0, 20))
            
            tk.Label(icon_frame,
                    text="📊",
                    font=("Segoe UI Emoji", 48),
                    bg=styles.COLOR_FONDO,
                    fg=self.system_color).pack()
            
            # LADO DERECHO: ESTADÍSTICAS EN COLUMNA
            numbers_frame = tk.Frame(stats_container, bg=styles.COLOR_FONDO)
            numbers_frame.pack(side=tk.LEFT, fill=tk.Y)
            
            # Productos (en lugar de Materiales)
            productos_frame = tk.Frame(numbers_frame, bg=styles.COLOR_FONDO)
            productos_frame.pack(anchor=tk.W, pady=(5, 15))
            
            tk.Label(productos_frame,
                    text="Productos:",
                    font=(styles.FUENTE_PRINCIPAL, 12),
                    bg=styles.COLOR_FONDO,
                    fg=styles.COLOR_TEXTO_OSCURO,
                    anchor="w").pack(side=tk.LEFT)
            
            tk.Label(productos_frame,
                    text=f"  {total_productos}",
                    font=(styles.FUENTE_PRINCIPAL, 14, styles.PESO_NEGRITA),
                    bg=styles.COLOR_FONDO,
                    fg=self.system_color,
                    anchor="w").pack(side=tk.LEFT)
            
            # Registros
            registros_frame = tk.Frame(numbers_frame, bg=styles.COLOR_FONDO)
            registros_frame.pack(anchor=tk.W, pady=(0, 5))
            
            tk.Label(registros_frame,
                    text="Registros: ",
                    font=(styles.FUENTE_PRINCIPAL, 12),
                    bg=styles.COLOR_FONDO,
                    fg=styles.COLOR_TEXTO_OSCURO,
                    anchor="w").pack(side=tk.LEFT)
            
            tk.Label(registros_frame,
                    text=f"  {total_registros}",
                    font=(styles.FUENTE_PRINCIPAL, 14, styles.PESO_NEGRITA),
                    bg=styles.COLOR_FONDO,
                    fg=self.system_color,
                    anchor="w").pack(side=tk.LEFT)
            
            # Frame para botones
            button_frame = tk.Frame(main_frame, bg=styles.COLOR_FONDO)
            button_frame.pack(fill=tk.X, pady=(10, 0))
            
            # Función para exportar REAL (específica de TIENDA)
            def do_export():
                confirm_window.destroy()
                
                try:
                    # Crear workbook
                    wb = openpyxl.Workbook()
                    
                    # HOJA 1: RESUMEN DE PRODUCTOS (código específico de TIENDA)
                    ws1 = wb.active
                    ws1.title = "Resumen Productos"
                    
                    # Título
                    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
                    ws1.merge_cells('A1:K1')
                    title_cell = ws1['A1']
                    title_cell.value = f"INVENTARIO TIENDA - Exportado el {fecha_actual}"
                    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # Encabezados
                    if data['resumen']:
                        headers = list(data['resumen'][0].keys())
                        for col_num, header in enumerate(headers, 1):
                            cell = ws1.cell(row=3, column=col_num, value=header)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        
                        # Datos
                        for row_num, row_data in enumerate(data['resumen'], 4):
                            for col_num, key in enumerate(headers, 1):
                                value = row_data[key]
                                
                                # Formato especial para precios (TIENDA tiene precios)
                                if "Precio" in key and value is not None:
                                    try:
                                        value = float(value)
                                        cell = ws1.cell(row=row_num, column=col_num, value=value)
                                        cell.number_format = '"₡"#,##0.00'
                                    except:
                                        cell = ws1.cell(row=row_num, column=col_num, value=value)
                                # Formato para números enteros
                                elif key in ["Stock Total", "Nivel Alarma", "Ubicaciones"] and value is not None:
                                    try:
                                        cell = ws1.cell(row=row_num, column=col_num, value=int(value))
                                        cell.number_format = '#,##0'
                                    except:
                                        cell = ws1.cell(row=row_num, column=col_num, value=value)
                                else:
                                    cell = ws1.cell(row=row_num, column=col_num, value=value)
                                
                                # Resaltar stock bajo
                                if key == "Stock Total" and value is not None:
                                    try:
                                        stock = int(value)
                                        alarma_col = headers.index("Nivel Alarma") + 1
                                        alarma_val = ws1.cell(row=row_num, column=alarma_col).value
                                        if alarma_val and stock < int(alarma_val):
                                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                    except:
                                        pass
                        
                        column_widths = {
                            'A': 35,  # Producto
                            'B': 20,  # Marca
                            'C': 20,  # Categoría
                            'D': 15,  # Precio Venta
                            'E': 15,  # Precio Compra
                            'F': 12,  # Color
                            'G': 10,  # Talla
                            'H': 15,  # Nivel Alarma
                            'I': 15,  # Stock Total
                            'J': 15,  # Ubicaciones
                            'K': 30,  # Edificios
                        }
                        
                        for col_letter, width in column_widths.items():
                            ws1.column_dimensions[col_letter].width = width
                    
                    # HOJA 2: DETALLE DE INVENTARIO
                    if data['detalle']:
                        ws2 = wb.create_sheet(title="Inventario Detallado")
                        
                        ws2.merge_cells('A1:H1')
                        title_cell = ws2['A1']
                        title_cell.value = f"DETALLE DE INVENTARIO POR EDIFICIO - {fecha_actual}"
                        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
                        title_cell.alignment = Alignment(horizontal='center', vertical='center')
                        title_cell.fill = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
                        
                        headers = list(data['detalle'][0].keys())
                        for col_num, header in enumerate(headers, 1):
                            cell = ws2.cell(row=3, column=col_num, value=header)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
                        
                        for row_num, row_data in enumerate(data['detalle'], 4):
                            for col_num, key in enumerate(headers, 1):
                                value = row_data[key]
                                
                                if key == "Cantidad" and value is not None:
                                    try:
                                        cell = ws2.cell(row=row_num, column=col_num, value=int(value))
                                        cell.number_format = '#,##0'
                                        if value == 0:
                                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                    except:
                                        cell = ws2.cell(row=row_num, column=col_num, value=value)
                                else:
                                    cell = ws2.cell(row=row_num, column=col_num, value=value)
                        
                        column_widths2 = {
                            'A': 35,  # Producto
                            'B': 20,  # Edificio
                            'C': 12,  # Cantidad
                            'D': 10,  # Estante
                            'E': 15,  # Lugar
                            'F': 20,  # Etiqueta
                            'G': 20,  # Marca
                            'H': 20,  # Categoría
                        }
                        
                        for col_letter, width in column_widths2.items():
                            ws2.column_dimensions[col_letter].width = width
                    
                    # GUARDAR ARCHIVO
                    export_dir = "exportaciones"
                    if not os.path.exists(export_dir):
                        os.makedirs(export_dir)
                    
                    fecha_nombre = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"{export_dir}/Inventario_TIENDA_{fecha_nombre}.xlsx"
                    
                    wb.save(filename)
                    
                    messagebox.showinfo(
                        "Exportación Exitosa", 
                        f"Archivo exportado correctamente:\n{filename}\n\n"
                        f"Total productos: {total_productos}\n"
                        f"Registros de inventario: {total_registros}"
                    )
                    
                    print(f"[TIENDA] Excel exportado: {filename}")
                    
                except Exception as e:
                    print(f"Error exportando Excel TIENDA: {e}")
                    messagebox.showerror(
                        "Error de Exportación", 
                        f"No se pudo exportar el archivo:\n{str(e)}"
                    )
            
            # Botón SÍ (Exportar)
            btn_export = tk.Button(button_frame,
                                text="Sí, Exportar",
                                font=(styles.FUENTE_PRINCIPAL, 11, styles.PESO_NEGRITA),
                                bg=styles.COLOR_EXITO,
                                fg=styles.COLOR_BLANCO,
                                relief=tk.FLAT,
                                bd=0,
                                cursor=styles.CURSOR_BOTON,
                                command=do_export,
                                padx=30,
                                pady=20)
            btn_export.pack(side=tk.LEFT, padx=(0, 25))
            
            # Botón NO (Cancelar)
            btn_cancel = tk.Button(button_frame,
                                text="No, Cancelar",
                                font=(styles.FUENTE_PRINCIPAL, 11, styles.PESO_NEGRITA),
                                bg=styles.COLOR_PELIGRO,
                                fg=styles.COLOR_BLANCO,
                                relief=tk.FLAT,
                                bd=0,
                                cursor=styles.CURSOR_BOTON,
                                command=confirm_window.destroy,
                                padx=30,
                                pady=20)
            btn_cancel.pack(side=tk.RIGHT)
            
            # Forzar actualización
            confirm_window.update_idletasks()
            
        except Exception as e:
            print(f"Error preparando exportación TIENDA: {e}")
            messagebox.showerror("Error", f"No se pudo preparar la exportación:\n{str(e)}")